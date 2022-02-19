import os
import json
import csv

from django.conf import settings
from openpyxl import load_workbook

from horeca_utils import constants


def get_state_context_default_value():
    return {'current_step': constants.TestingSessionSteps.DATA_PROCESSING_POLICY.value}


def get_values_by_gender(values, gender):
    '''Возвращает список значений в зависимости от пола кандидата'''
    if gender == constants.Genders.MALE.value:
        values = values.exclude(scale__name=constants.MMPIScales.MF_M.value)
    else:
        values = values.exclude(scale__name=constants.MMPIScales.MM.value)

    return values


class TestingDataParser:
    source_dir = os.path.join(settings.BASE_DIR, 'candidate_tests', 'testing_data')

    def parse_MMPI_questions_file(self, file_name: str) -> list:
        file_name = os.path.join(self.source_dir, file_name)
        with open(file_name) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            questions = []

            for row in csv_reader:
                question = {
                    'id': row[0],
                    'question': row[1],
                }
                questions.append(question)

            return questions

    def parse_MMPI_scales_description(self, file_name='MMPI_scales_description.xlsx'):
        descriptions = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Описания')

        for row in range(2, sheet.max_row+1):
            description = {
                'verbose_name': sheet[f'A{row}'].value,
                'from_point': sheet[f'D{row}'].value,
                'to_point': sheet[f'E{row}'].value,
                'description': sheet[f'F{row}'].value,
            }
            description['description_detail'] = sheet[f'G{row}'].value if sheet[f'G{row}'].value else ''
            descriptions.append(description)

        return descriptions

    def parse_MMPI_motivators_destructors(self, sheet_name, file_name='MMPI_motivators_destructors.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name(sheet_name)
        title = 'motivator' if sheet_name == 'Motivators' else 'destructor'

        for row in range(1, sheet.max_row+1):
            value = {
                'scale': sheet[f'B{row}'].value,
                'from_point': sheet[f'C{row}'].value,
                'to_point': sheet[f'D{row}'].value,
                title: sheet[f'E{row}'].value,
                'description': sheet[f'F{row}'].value,
                'recommendations': sheet[f'G{row}'].value if sheet[f'G{row}'].value else '',
            }
            values.append(value)

        return values

    def parse_MMPI_risk_factors(self, sheet_name, file_name='MMPI_risk_factors.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name(sheet_name)

        for row in range(2, sheet.max_row+1):
            value = {
                'scale': sheet[f'B{row}'].value,
                'from_point': sheet[f'C{row}'].value,
                'to_point': sheet[f'D{row}'].value,
                'factor': sheet[f'E{row}'].value,
                'description': sheet[f'F{row}'].value,
                'is_attantion_factor': True if sheet_name == 'Attantion factors' else False,
            }
            values.append(value)

        return values

    def parse_MMPI_competences_category(self, file_name='MMPI_competences.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Competences')

        for row in range(2, sheet.max_row+1):
            values.append(sheet[f'F{row}'].value)

        return set(values)

    def parse_MMPI_competences(self, file_name='MMPI_competences.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Competences')

        for row in range(2, sheet.max_row+1):
            value = {
                'name': sheet[f'D{row}'].value,
                'description': sheet[f'E{row}'].value,
                'category': sheet[f'F{row}'].value,
            }
            values.append(value)

        values = [dict(t) for t in {tuple(v.items()) for v in values}]

        return values

    def parse_MMPI_competences_scales(self, file_name='MMPI_competences.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Competences')

        for row in range(2, sheet.max_row+1):
            value = {
                'scale': sheet[f'A{row}'].value,
                'from_point': sheet[f'B{row}'].value,
                'to_point': sheet[f'C{row}'].value,
                'competence': sheet[f'D{row}'].value,
            }
            values.append(value)

        return values

    def parse_MMPI_team_roles(self, file_name='MMPI_team_roles.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Roles')

        for row in range(2, sheet.max_row+1):
            value = {
                'name': sheet[f'A{row}'].value,
                'description': sheet[f'B{row}'].value,
            }
            values.append(value)

        return values

    def parse_MMPI_team_role_scales(self, file_name='MMPI_team_roles.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Scales')

        for row in sheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if cell.value is not None:
                    value = {
                        'role': sheet.cell(row=cell.row, column=1).value,
                        'scale': sheet.cell(row=1, column=cell.column).value,
                        'points': cell.value,
                    }
                    values.append(value)
        return values

    def parse_MMPI_stress_tolerance(self, file_name='MMPI_stress_tolerance.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Sheet2')

        for row in range(1, sheet.max_row+1):
            value = {
                'name': sheet[f'A{row}'].value,
                'description': sheet[f'B{row}'].value,
            }
            values.append(value)

        return values

    def parse_MMPI_stress_tolerance_scales(self, file_name='MMPI_stress_tolerance.xlsx'):
        values = {
            constants.MMPIStressTolerances.HIGHT.value: [],
            constants.MMPIStressTolerances.MEDIUM.value: [],
            constants.MMPIStressTolerances.LOW.value: [],
        }
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Sheet3')

        for row in range(2, sheet.max_row+1):
            if sheet[f'A{row}'].value == 'Сочетания шкал ЗА':
                stress_tolerance = sheet[f'E{row}'].value
                scales_combine = []
                for r in range(row, sheet.max_row+1):
                    scales_combine.append({
                        'scale': sheet[f'B{r}'].value,
                        'from_point': sheet[f'C{r}'].value,
                        'to_point': sheet[f'D{r}'].value,
                    })
                    if sheet[f'A{r+1}'].value == 'Сочетания шкал ЗА':
                        values[stress_tolerance].append(scales_combine)
                        break

        return values

    def parse_MMPI_leadership_styles(self, file_name='MMPI_leadership_styles.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Sheet2')

        for row in range(2, sheet.max_row+1):
            value = {
                'name': sheet[f'A{row}'].value,
                'description': sheet[f'B{row}'].value,
            }
            values.append(value)

        return values

    def parse_MMPI_leadership_style_scales(self, file_name='MMPI_leadership_styles.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Sheet3')

        for row in range(2, sheet.max_row+1):
            value = {
                'scale': sheet[f'A{row}'].value,
                'from_point': sheet[f'B{row}'].value,
                'to_point': sheet[f'C{row}'].value,
                'style': sheet[f'D{row}'].value,
            }
            values.append(value)

        return values

    def parse_logic_test_file(self, file_name):
        questions = []
        max_row_range = 24 if file_name == 'numeric_logic_questions.xlsx' else 29
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Вопросы')

        for index, row in enumerate(range(2, max_row_range)):
            question = {
                'id': index + 1,
                'question': sheet[f'A{row}'].value,
                'answers': [],
            }
            for column in range(ord('B'), ord('L')+1, 2):
                answer = sheet[f'{chr(column)}{row}'].value
                point = sheet[f'{chr(column+1)}{row}'].value
                if answer is None:
                    break
                question['answers'].append({
                    'answer': answer,
                    'point': point,
                })

            questions.append(question)

        return questions

    def parse_prof_interests_scales(self, file_name='prof_interests.xlsx'):
        scales = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Описания')

        row = 1
        while True:
            number = sheet[f'A{row}'].value
            if number is not None:
                scale = {
                    'number': number,
                    'name': sheet[f'B{row}'].value,
                    'description': sheet[f'C{row}'].value,
                }
                scales.append(scale)
                row += 1
            else:
                break

        return scales

    def parse_prof_interests_questions(self, file_name='prof_interests.xlsx'):
        questions = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Вопросы')

        row = 2
        while True:
            id = sheet[f'A{row}'].value
            if id is not None:
                question = {
                    'id': id,
                    'answer': sheet[f'B{row}'].value,
                    'scale': sheet[f'C{row}'].value
                }
                questions.append(question)
                row += 1
            else:
                break

        return questions

    def parse_test_answers_file(self, candidate_name, file_name='test_answers.xlsx'):
        values = []
        file_name = os.path.join(self.source_dir, file_name)
        wb = load_workbook(file_name)
        sheet = wb.get_sheet_by_name('Answers')
        start_column = None
        candidate_row = None

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == candidate_name:
                    candidate_row = cell.row
                if cell.value == constants.TestTypes.MMPI.value:
                    start_column = cell.column

        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=start_column, max_col=sheet.max_column-1):
            for cell in row:
                answer = sheet.cell(candidate_row, cell.col_idx).value
                answer = constants.MMPIAnswers.YES.value if answer == 'ДА' else constants.MMPIAnswers.NO.value
                answer = {
                    'question': cell.value,
                    'answer': answer,
                }
                values.append(answer)

        return values

    def parse_json_file(self, file_name: str):
        file_name = os.path.join(self.source_dir, file_name)
        with open(file_name) as json_file:
            return json.load(json_file)


testing_data_parser = TestingDataParser()
